from django.shortcuts import render
from django.http import HttpResponse
from django.db import models

from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from .models import Auditoria
from core.decoradores import requerir_rol


# ============================================================
# OBTENER AUDITORÍAS FILTRADAS
# ============================================================

def obtener_auditorias_filtradas(request):

    modulo = request.GET.get("modulo", "").strip()
    usuario = request.GET.get("usuario", "").strip()
    accion = request.GET.get("accion", "").strip()

    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()

    hoy = date.today()

    datos = Auditoria.objects.all().order_by("-fecha")

    # ========================================================
    # FILTRO POR MÓDULO
    # ========================================================

    if modulo:
        datos = datos.filter(
            modulo=modulo
        )

    # ========================================================
    # FILTRO POR USUARIO
    # ========================================================

    if usuario:
        datos = datos.filter(
            usuario__icontains=usuario
        )

    # ========================================================
    # FILTRO POR ACCIÓN
    # ========================================================

    if accion:
        datos = datos.filter(
            accion__icontains=accion
        )

    # ========================================================
    # FILTRO POR FECHAS
    # ========================================================

    try:

        # ----------------------------------------------------
        # FECHA DESDE
        # ----------------------------------------------------

        if fecha_desde:

            d = datetime.fromisoformat(
                fecha_desde
            )

            # No permitir fechas futuras
            if d.date() > hoy:

                fecha_desde = hoy.isoformat()

                d = datetime.fromisoformat(
                    fecha_desde
                )

            datos = datos.filter(
                fecha__gte=d
            )

        # ----------------------------------------------------
        # FECHA HASTA
        # ----------------------------------------------------

        if fecha_hasta:

            h = datetime.fromisoformat(
                fecha_hasta
            )

            # No permitir fechas futuras
            if h.date() > hoy:

                fecha_hasta = hoy.isoformat()

                h = datetime.fromisoformat(
                    fecha_hasta
                )

            # Si solamente viene la fecha,
            # incluir todo el día
            if (
                h.hour == 0
                and h.minute == 0
                and h.second == 0
            ):
                h = h.replace(
                    hour=23,
                    minute=59,
                    second=59
                )

            datos = datos.filter(
                fecha__lte=h
            )

    except (ValueError, TypeError):

        # Si la fecha tiene un formato inválido,
        # simplemente ignoramos el filtro.

        pass

    return datos


# ============================================================
# AUDITORÍA - LISTADO
# ============================================================

@requerir_rol(["SuperAdmin", "Admin"])
def auditoria(request):

    datos = obtener_auditorias_filtradas(
        request
    )

    modulo = request.GET.get(
        "modulo",
        ""
    ).strip()

    usuario = request.GET.get(
        "usuario",
        ""
    ).strip()

    accion = request.GET.get(
        "accion",
        ""
    ).strip()

    fecha_desde = request.GET.get(
        "fecha_desde",
        ""
    ).strip()

    fecha_hasta = request.GET.get(
        "fecha_hasta",
        ""
    ).strip()

    hoy = date.today()

    modulos = Auditoria.MODULOS

    return render(
        request,
        "auditoria/ver_registros.html",
        {
            "datos": datos,

            "modulos": modulos,

            "modulo_actual": modulo,

            "usuario_filtro": usuario,

            "accion_filtro": accion,

            "fecha_desde": fecha_desde,

            "fecha_hasta": fecha_hasta,

            "hoy": hoy.isoformat(),
        }
    )


# ============================================================
# EXPORTAR AUDITORÍA A EXCEL
# ============================================================

@requerir_rol(["SuperAdmin", "Admin"])
def exportar_excel(request):

    datos = obtener_auditorias_filtradas(
        request
    )

    # Crear archivo Excel
    workbook = Workbook()

    hoja = workbook.active

    hoja.title = "Auditoría"

    # ========================================================
    # TÍTULO
    # ========================================================

    hoja["A1"] = "SIGIF"
    hoja["A1"].font = Font(
        bold=True,
        size=16
    )

    hoja["A2"] = (
        "Informe de Auditoría"
    )

    hoja["A2"].font = Font(
        bold=True,
        size=14
    )

    # ========================================================
    # ENCABEZADOS
    # ========================================================

    encabezados = [
        "Fecha",
        "Hora",
        "Usuario",
        "Acción",
        "Módulo"
    ]

    fila_encabezado = 4

    for columna, encabezado in enumerate(
        encabezados,
        start=1
    ):

        celda = hoja.cell(
            row=fila_encabezado,
            column=columna
        )

        celda.value = encabezado

        celda.font = Font(
            bold=True
        )

        celda.alignment = Alignment(
            horizontal="center"
        )

    # ========================================================
    # DATOS
    # ========================================================

    fila = 5

    for registro in datos:

        fecha_local = registro.fecha

        hoja.cell(
            row=fila,
            column=1
        ).value = fecha_local.strftime(
            "%d/%m/%Y"
        )

        hoja.cell(
            row=fila,
            column=2
        ).value = fecha_local.strftime(
            "%H:%M:%S"
        )

        hoja.cell(
            row=fila,
            column=3
        ).value = registro.usuario

        hoja.cell(
            row=fila,
            column=4
        ).value = registro.accion

        hoja.cell(
            row=fila,
            column=5
        ).value = registro.modulo

        fila += 1

    # ========================================================
    # AJUSTAR ANCHO DE COLUMNAS
    # ========================================================

    hoja.column_dimensions["A"].width = 15
    hoja.column_dimensions["B"].width = 12
    hoja.column_dimensions["C"].width = 25
    hoja.column_dimensions["D"].width = 60
    hoja.column_dimensions["E"].width = 20

    # ========================================================
    # CONGELAR ENCABEZADOS
    # ========================================================

    hoja.freeze_panes = "A5"

    # ========================================================
    # RESPUESTA
    # ========================================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="informe_auditoria.xlsx"'
    )

    workbook.save(
        response
    )

    return response


# ============================================================
# EXPORTAR AUDITORÍA A PDF
# ============================================================

@requerir_rol(["SuperAdmin", "Admin"])
def exportar_pdf(request):

    datos = obtener_auditorias_filtradas(
        request
    )

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="informe_auditoria.pdf"'
    )

    # ========================================================
    # DOCUMENTO
    # ========================================================

    documento = SimpleDocTemplate(
        response,
        pagesize=landscape(letter),

        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    elementos = []

    estilos = getSampleStyleSheet()

    # ========================================================
    # TÍTULO
    # ========================================================

    estilo_titulo = estilos["Title"]

    estilo_titulo.fontSize = 18

    elementos.append(
        Paragraph(
            "SIGIF",
            estilo_titulo
        )
    )

    elementos.append(
        Paragraph(
            "Informe de Auditoría",
            estilos["Heading2"]
        )
    )

    elementos.append(
        Spacer(
            1,
            15
        )
    )

    # ========================================================
    # TABLA
    # ========================================================

    tabla_datos = [
        [
            "Fecha",
            "Hora",
            "Usuario",
            "Acción",
            "Módulo"
        ]
    ]

    for registro in datos:

        fecha_local = registro.fecha

        tabla_datos.append(
            [
                fecha_local.strftime(
                    "%d/%m/%Y"
                ),

                fecha_local.strftime(
                    "%H:%M:%S"
                ),

                registro.usuario,

                registro.accion,

                registro.modulo
            ]
        )

    # ========================================================
    # CREAR TABLA
    # ========================================================

    tabla = Table(
        tabla_datos,

        colWidths=[
            70,
            60,
            100,
            350,
            100
        ],

        repeatRows=1
    )

    tabla.setStyle(
        TableStyle(
            [

                # ------------------------------------------------
                # ENCABEZADO
                # ------------------------------------------------

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor(
                        "#1E2A44"
                    )
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),

                (
                    "ALIGN",
                    (0, 0),
                    (-1, 0),
                    "CENTER"
                ),

                # ------------------------------------------------
                # CUERPO
                # ------------------------------------------------

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE"
                ),

                (
                    "FONTSIZE",
                    (0, 1),
                    (-1, -1),
                    8
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),
            ]
        )
    )

    elementos.append(
        tabla
    )

    elementos.append(
        Spacer(
            1,
            15
        )
    )

    elementos.append(
        Paragraph(
            "SIGIF - Sistema Integral de Gestión "
            "de Inventarios y Facturación",
            estilos["Normal"]
        )
    )

    # ========================================================
    # GENERAR PDF
    # ========================================================

    documento.build(
        elementos
    )

    return response